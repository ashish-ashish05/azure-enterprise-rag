from io import BytesIO

from openpyxl import load_workbook

from enterprise_rag.domain.models import Document
from enterprise_rag.ingestion.loaders.base import DocumentLoader


class XlsxDocumentLoader(DocumentLoader):
    """Extract spreadsheet content from XLSX documents."""

    def load(
        self,
        content: bytes,
        *,
        document_id: str,
        document_family_id: str,
        source: str,
    ) -> Document:
        workbook = load_workbook(
            filename=BytesIO(content),
            read_only=True,
            data_only=True,
        )

        sheet_count = len(workbook.sheetnames)
        sections: list[str] = []

        for worksheet in workbook.worksheets:
            rows: list[str] = []

            for row in worksheet.iter_rows(
                values_only=True
            ):
                values = [
                    str(value).strip()
                    for value in row
                    if value is not None
                ]

                if values:
                    rows.append(
                        " | ".join(values)
                    )

            if rows:
                sections.append(
                    f"Sheet: {worksheet.title}\n"
                    + "\n".join(rows)
                )

        workbook.close()

        return Document(
            document_id=document_id,
            document_family_id=document_family_id,
            source=source,
            content="\n\n".join(sections),
            content_type=(
                "application/vnd.openxmlformats-officedocument"
                ".spreadsheetml.sheet"
            ),
            metadata={
                "sheet_count": sheet_count,
            },
        )